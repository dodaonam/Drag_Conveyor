"""Persistent Excel log for confirmed drag-conveyor inspections."""
from __future__ import annotations

import tempfile
import threading
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

LOG_FILENAME = "Nhat_ky_kiem_tra.xlsx"
SHEET_NAME = "Nhật ký kiểm tra"
META_SHEET_NAME = "_drag_conveyor_meta"
META_MARKER = "drag_conveyor_excel_log_v1"

HEADERS = (
    "Mã job",
    "Ngày kiểm tra",
    "Người kiểm tra",
    "Tên băng chuyền",
    "Tổng số thanh",
    "Tổng số thanh lỗi",
    "Tổng số thanh bình thường",
    "Số thanh lệch trái",
    "Số thanh lệch phải",
    "Số thanh lệch cả 2",
    "Số thanh gãy",
    "Tên file PDF",
)

_ICT = timezone(timedelta(hours=7))
_WRITE_LOCK = threading.Lock()


class ExcelLogError(Exception):
    """Raised when the inspection log cannot be safely updated."""


def save_inspection_log(
    *,
    report_data: dict,
    job_id: str,
    inspected_at_iso: str,
    inspector_name: str,
    conveyor_name: str,
    pdf_filename: str,
    reports_dir: str | Path,
) -> str:
    """Create or upsert one confirmed inspection row and return the workbook name."""
    reports_path = Path(reports_dir)
    reports_path.mkdir(parents=True, exist_ok=True)

    with _WRITE_LOCK:
        target = _resolve_log_path(reports_path)
        workbook, sheet = _load_or_create_workbook(target)
        values = _row_values(
            report_data=report_data,
            job_id=job_id,
            inspected_at_iso=inspected_at_iso,
            inspector_name=inspector_name,
            conveyor_name=conveyor_name,
            pdf_filename=pdf_filename,
        )
        _upsert_row(sheet, values)
        _save_atomically(workbook, target)
        return target.name


def _resolve_log_path(reports_dir: Path) -> Path:
    default_path = reports_dir / LOG_FILENAME
    candidates: list[Path] = []
    for path in reports_dir.glob("*.xlsx"):
        try:
            if _is_log_workbook(path):
                candidates.append(path)
        except Exception as exc:  # noqa: BLE001
            if path == default_path:
                raise ExcelLogError(f"Không đọc được file Excel mặc định: {exc}") from exc

    if default_path in candidates:
        return default_path
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise ExcelLogError(f"Tìm thấy nhiều file nhật ký Excel: {names}")
    if default_path.exists():
        raise ExcelLogError(
            f"{LOG_FILENAME} đã tồn tại nhưng không phải file nhật ký của Drag Conveyor"
        )
    return default_path


def _is_log_workbook(path: Path) -> bool:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return (
            META_SHEET_NAME in workbook.sheetnames
            and workbook[META_SHEET_NAME]["A1"].value == META_MARKER
        )
    finally:
        workbook.close()


def _load_or_create_workbook(target: Path):
    if target.exists():
        try:
            workbook = load_workbook(target)
        except Exception as exc:  # noqa: BLE001
            raise ExcelLogError(f"Không mở được file Excel: {exc}") from exc
        if not _is_current_log(workbook):
            workbook.close()
            raise ExcelLogError("File Excel không đúng định dạng nhật ký Drag Conveyor")
        return workbook, workbook[SHEET_NAME]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:L1"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
    sheet.column_dimensions["A"].hidden = True
    for column, width in {
        "B": 20, "C": 24, "D": 24, "E": 16, "F": 18, "G": 23,
        "H": 20, "I": 20, "J": 22, "K": 15, "L": 34,
    }.items():
        sheet.column_dimensions[column].width = width

    meta = workbook.create_sheet(META_SHEET_NAME)
    meta["A1"] = META_MARKER
    meta.sheet_state = "hidden"
    return workbook, sheet


def _is_current_log(workbook) -> bool:
    if META_SHEET_NAME not in workbook.sheetnames or SHEET_NAME not in workbook.sheetnames:
        return False
    if workbook[META_SHEET_NAME]["A1"].value != META_MARKER:
        return False
    headers = [cell.value for cell in workbook[SHEET_NAME][1]]
    return all(header in headers for header in HEADERS)


def _row_values(
    *,
    report_data: dict,
    job_id: str,
    inspected_at_iso: str,
    inspector_name: str,
    conveyor_name: str,
    pdf_filename: str,
) -> dict[str, object]:
    defects_by_type = report_data["defects_by_type"]
    defect_count = int(report_data["defect_count"])
    total_bars = int(report_data["total_bars"])
    return {
        "Mã job": job_id,
        "Ngày kiểm tra": _to_ict_naive(inspected_at_iso),
        "Người kiểm tra": inspector_name,
        "Tên băng chuyền": conveyor_name,
        "Tổng số thanh": total_bars,
        "Tổng số thanh lỗi": defect_count,
        "Tổng số thanh bình thường": total_bars - defect_count,
        "Số thanh lệch trái": len(defects_by_type["bent_left"]),
        "Số thanh lệch phải": len(defects_by_type["bent_right"]),
        "Số thanh lệch cả 2": len(defects_by_type["bent_both"]),
        "Số thanh gãy": len(defects_by_type["broken"]),
        "Tên file PDF": pdf_filename,
    }


def _to_ict_naive(value: str) -> datetime:
    parsed = datetime.fromisoformat(value)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(_ICT).replace(tzinfo=None)


def _upsert_row(sheet, values: dict[str, object]) -> None:
    header_columns = {str(cell.value): cell.column for cell in sheet[1]}
    job_column = header_columns["Mã job"]
    row_index = next(
        (row for row in range(2, sheet.max_row + 1) if sheet.cell(row, job_column).value == values["Mã job"]),
        sheet.max_row + 1,
    )
    for header, value in values.items():
        cell = sheet.cell(row_index, header_columns[header])
        cell.value = value
        if header == "Ngày kiểm tra":
            cell.number_format = "dd/mm/yyyy hh:mm"


def _save_atomically(workbook, target: Path) -> None:
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            suffix=".xlsx",
            prefix=f".{target.stem}.",
            dir=target.parent,
            delete=False,
        ) as handle:
            temp_path = Path(handle.name)
        workbook.save(temp_path)
        temp_path.replace(target)
    except PermissionError as exc:
        raise ExcelLogError("Hãy đóng file Excel nếu đang mở rồi thử lại") from exc
    except OSError as exc:
        raise ExcelLogError(f"Không thể ghi file Excel: {exc}") from exc
    finally:
        workbook.close()
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
