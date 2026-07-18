from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SERVER_DIR = ROOT / "server"
if str(SERVER_DIR) not in sys.path:
    sys.path.insert(0, str(SERVER_DIR))

import excel_log  # noqa: E402


def _report_data() -> dict:
    return {
        "total_bars": 10,
        "defect_count": 4,
        "defects_by_type": {
            "bent_left": [{}, {}],
            "bent_right": [{}],
            "bent_both": [],
            "broken": [{}],
        },
    }


class ExcelLogTests(unittest.TestCase):
    def _save(self, directory: Path, *, job_id: str = "job-1", pdf_filename: str = "job-1.pdf") -> str:
        return excel_log.save_inspection_log(
            report_data=_report_data(),
            job_id=job_id,
            inspected_at_iso="2026-07-18T03:20:00+00:00",
            inspector_name="Nguyễn Văn A",
            conveyor_name="Băng 1",
            pdf_filename=pdf_filename,
            reports_dir=directory,
        )

    def test_creates_log_with_expected_counts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            directory = Path(tmp)
            filename = self._save(directory)
            workbook = load_workbook(directory / filename, data_only=True)
            sheet = workbook[excel_log.SHEET_NAME]
            self.assertEqual(sheet.max_row, 2)
            self.assertEqual([cell.value for cell in sheet[2]], [
                "job-1", sheet["B2"].value, "Nguyễn Văn A", "Băng 1", 10, 4, 6, 2, 1, 0, 1, "job-1.pdf",
            ])
            self.assertTrue(workbook[excel_log.META_SHEET_NAME].sheet_state == "hidden")
            workbook.close()

    def test_upserts_same_job_and_discovers_renamed_log(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            directory = Path(tmp)
            original = directory / self._save(directory)
            renamed = directory / "Nhat_ky_thang_7.xlsx"
            original.rename(renamed)

            filename = self._save(directory, pdf_filename="job-1-v2.pdf")
            self.assertEqual(filename, renamed.name)
            workbook = load_workbook(renamed, data_only=True)
            sheet = workbook[excel_log.SHEET_NAME]
            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(sheet["L2"].value, "job-1-v2.pdf")
            workbook.close()

    def test_adds_another_job_as_a_new_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            directory = Path(tmp)
            self._save(directory)
            self._save(directory, job_id="job-2", pdf_filename="job-2.pdf")
            workbook = load_workbook(directory / excel_log.LOG_FILENAME, data_only=True)
            self.assertEqual(workbook[excel_log.SHEET_NAME].max_row, 3)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
